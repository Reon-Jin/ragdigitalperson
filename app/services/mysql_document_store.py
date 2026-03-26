from __future__ import annotations

import hashlib
import json
import re
import shutil
import uuid
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from html import unescape
from pathlib import Path
from typing import Any, Dict, List, Sequence

import numpy as np
from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

from app.config import Settings
from app.storage.database import Database


SUPPORTED_SUFFIXES = {".txt", ".md", ".pdf", ".docx", ".csv", ".json", ".html", ".htm", ".xlsx"}
NOISE_PATTERNS = (
    r"^page\s+\d+$",
    r"^\d+$",
    r"^doi[:\s]",
    r"^(received|accepted|published online)",
    r"^(copyright|preprint)",
    r"^vol\.",
    r"^figure\s+\d+",
    r"^table\s+\d+",
)


@dataclass
class SearchResult:
    doc_id: str
    filename: str
    category: str
    title: str
    section_id: str
    section_title: str
    chunk_id: str
    chunk_index: int
    chunk_title: str
    page_start: int | None
    page_end: int | None
    chunk_kind: str
    score: float
    text: str


class DocumentStore:
    def __init__(self, settings: Settings) -> None:
        self.settings = settings
        self.db = Database(settings.database_url, settings.app_db_path)
        self.db.init_schema()
        self.docs_path = self.settings.data_dir / "documents.json"
        self.sections_path = self.settings.data_dir / "sections.json"
        self.chunks_path = self.settings.data_dir / "chunks.json"
        self.pages_path = self.settings.data_dir / "pages.json"
        self._bootstrap_legacy_json_if_needed()
        self._refresh_from_db()

    def user_upload_dir(self, user_id: str) -> Path:
        path = self.settings.uploads_dir / user_id
        path.mkdir(parents=True, exist_ok=True)
        return path

    def resolve_storage_path(self, doc_id: str, suffix: str, user_id: str) -> Path:
        return self.user_upload_dir(user_id) / f"{doc_id}{suffix}"

    def _refresh_from_db(self) -> None:
        self.docs = []
        self.sections = []
        self.chunks = []
        self.pages = []
        self.doc_embeddings_by_id: dict[str, np.ndarray] = {}
        self.section_embeddings_by_id: dict[str, np.ndarray] = {}
        self.chunk_embeddings_by_id: dict[str, np.ndarray] = {}

        for row in self.db.fetchall("SELECT * FROM documents ORDER BY uploaded_at DESC"):
            payload = dict(row)
            doc_id = str(payload.get("doc_id") or "")
            payload["headings"] = self._loads(payload.pop("headings_json", "[]"), [])
            payload["keywords"] = self._loads(payload.pop("keywords_json", "[]"), [])
            self.doc_embeddings_by_id[doc_id] = self._to_vector(payload.pop("embedding_json", "[]"))
            self.docs.append(payload)

        for row in self.db.fetchall("SELECT * FROM document_sections ORDER BY order_index ASC"):
            payload = dict(row)
            section_id = str(payload.get("section_id") or "")
            payload["order"] = int(payload.pop("order_index", 0) or 0)
            self.section_embeddings_by_id[section_id] = self._to_vector(payload.pop("embedding_json", "[]"))
            self.sections.append(payload)

        for row in self.db.fetchall("SELECT * FROM document_chunks ORDER BY chunk_index ASC"):
            payload = dict(row)
            chunk_id = str(payload.get("chunk_id") or "")
            self.chunk_embeddings_by_id[chunk_id] = self._to_vector(payload.pop("embedding_json", "[]"))
            self.chunks.append(payload)

        self.pages = [dict(row) for row in self.db.fetchall("SELECT * FROM document_pages ORDER BY doc_id, page_number ASC")]
        self._refresh_maps()
        self._refresh_search_index()

    def _loads(self, value: str | None, default: Any) -> Any:
        if not value:
            return default
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            return default

    def _to_vector(self, value: str | list[float] | np.ndarray | None) -> np.ndarray:
        if isinstance(value, np.ndarray):
            return value.astype(np.float32)
        if isinstance(value, list):
            return np.asarray(value, dtype=np.float32)
        if not value:
            return np.zeros(self.settings.embedding_dimensions, dtype=np.float32)
        return np.asarray(self._loads(str(value), []), dtype=np.float32)

    def _encode_text(self, text: str) -> np.ndarray:
        cleaned = re.sub(r"\s+", " ", str(text or "")).strip()
        if not cleaned:
            return np.zeros(self.settings.embedding_dimensions, dtype=np.float32)
        vector = np.zeros(self.settings.embedding_dimensions, dtype=np.float32)
        words = cleaned.split()
        for word in words or [cleaned]:
            padded = f" {word} "
            for n in range(2, 5):
                if len(padded) < n:
                    continue
                for index in range(len(padded) - n + 1):
                    gram = padded[index:index + n]
                    digest = hashlib.md5(gram.encode('utf-8')).hexdigest()
                    bucket = int(digest[:8], 16) % self.settings.embedding_dimensions
                    vector[bucket] += 1.0
        norm = float(np.linalg.norm(vector))
        if norm > 0:
            vector /= norm
        return vector

    def _vector_json(self, vector: np.ndarray) -> str:
        return json.dumps([round(float(item), 6) for item in vector.tolist()], ensure_ascii=False)

    def _read_document_payload(self, file_path: Path) -> Dict[str, Any]:
        suffix = file_path.suffix.lower()
        if suffix in {".txt", ".md", ".csv"}:
            text = file_path.read_text(encoding="utf-8", errors="ignore")
            return {"text": text, "pages": [{"page_number": 1, "text": text}]}
        if suffix == ".xlsx":
            workbook = load_workbook(filename=str(file_path), data_only=True, read_only=True)
            blocks: list[str] = []
            for worksheet in workbook.worksheets:
                rows: list[str] = []
                for row in worksheet.iter_rows(values_only=True):
                    cells = [str(cell).strip() for cell in row if cell not in (None, "")]
                    if cells:
                        rows.append(" | ".join(cells))
                if rows:
                    blocks.append(f"# {worksheet.title}\n" + "\n".join(rows))
            text = "\n\n".join(blocks).strip()
            return {"text": text, "pages": [{"page_number": 1, "text": text}]}
        if suffix == ".json":
            raw = file_path.read_text(encoding="utf-8", errors="ignore")
            try:
                text = json.dumps(json.loads(raw), ensure_ascii=False, indent=2)
            except json.JSONDecodeError:
                text = raw
            return {"text": text, "pages": [{"page_number": 1, "text": text}]}
        if suffix in {".html", ".htm"}:
            raw = file_path.read_text(encoding="utf-8", errors="ignore")
            stripped = re.sub(r"<script[\s\S]*?</script>", " ", raw, flags=re.IGNORECASE)
            stripped = re.sub(r"<style[\s\S]*?</style>", " ", stripped, flags=re.IGNORECASE)
            text = unescape(re.sub(r"<[^>]+>", " ", stripped))
            return {"text": text, "pages": [{"page_number": 1, "text": text}]}
        if suffix == ".pdf":
            reader = PdfReader(str(file_path))
            pages = [{"page_number": index + 1, "text": page.extract_text() or ""} for index, page in enumerate(reader.pages)]
            return {"text": "\n\n".join(page["text"] for page in pages), "pages": pages}
        if suffix == ".docx":
            document = Document(str(file_path))
            text = "\n".join(paragraph.text for paragraph in document.paragraphs)
            return {"text": text, "pages": [{"page_number": 1, "text": text}]}
        raise ValueError(f"Unsupported file type: {suffix}")


    def _refresh_maps(self) -> None:
        self.docs_by_id = {item["doc_id"]: item for item in self.docs}
        self.sections_by_id = {item["section_id"]: item for item in self.sections}
        self.chunks_by_id = {item["chunk_id"]: item for item in self.chunks}
        self.pages_by_doc_id = {
            doc_id: sorted([page for page in self.pages if page["doc_id"] == doc_id], key=lambda item: item.get("page_number", 0))
            for doc_id in self.docs_by_id
        }
        self.doc_index_by_id = {item["doc_id"]: index for index, item in enumerate(self.docs)}
        self.section_index_by_id = {item["section_id"]: index for index, item in enumerate(self.sections)}
        self.chunk_index_by_id = {item["chunk_id"]: index for index, item in enumerate(self.chunks)}

    def _tokenize_search_text(self, text: str) -> list[str]:
        normalized = re.sub(r"\s+", " ", str(text or "")).lower().strip()
        if not normalized:
            return []
        tokens: list[str] = []
        tokens.extend(re.findall(r"[a-z0-9]{2,}", normalized))
        for fragment in re.findall(r"[\u4e00-\u9fff]{2,}", normalized):
            if len(fragment) <= 4:
                tokens.append(fragment)
            for width in (2, 3):
                if len(fragment) < width:
                    continue
                tokens.extend(fragment[index:index + width] for index in range(len(fragment) - width + 1))
        return tokens[:256]

    def _compute_token_idf(self, counters: list[Counter[str]]) -> dict[str, float]:
        if not counters:
            return {}
        document_frequency: Counter[str] = Counter()
        for counter in counters:
            document_frequency.update(counter.keys())
        total = len(counters)
        return {
            token: 1.0 + np.log((1.0 + total) / (1.0 + freq))
            for token, freq in document_frequency.items()
        }

    def _refresh_search_index(self) -> None:
        self.doc_ids = [item["doc_id"] for item in self.docs]
        self.chunk_ids = [item["chunk_id"] for item in self.chunks]
        self.doc_matrix = (
            np.vstack([self.doc_embeddings_by_id.get(doc_id, np.zeros(self.settings.embedding_dimensions, dtype=np.float32)) for doc_id in self.doc_ids]).astype(np.float32)
            if self.doc_ids
            else np.zeros((0, self.settings.embedding_dimensions), dtype=np.float32)
        )
        self.chunk_matrix = (
            np.vstack([self.chunk_embeddings_by_id.get(chunk_id, np.zeros(self.settings.embedding_dimensions, dtype=np.float32)) for chunk_id in self.chunk_ids]).astype(np.float32)
            if self.chunk_ids
            else np.zeros((0, self.settings.embedding_dimensions), dtype=np.float32)
        )
        self.doc_tokens_by_id = {
            doc_id: Counter(self._tokenize_search_text(self._doc_search_text(self.docs_by_id[doc_id])))
            for doc_id in self.doc_ids
        }
        self.chunk_tokens_by_id = {
            chunk_id: Counter(self._tokenize_search_text(self._chunk_search_text(self.chunks_by_id[chunk_id])))
            for chunk_id in self.chunk_ids
        }
        self.doc_token_idf = self._compute_token_idf(list(self.doc_tokens_by_id.values()))
        self.chunk_token_idf = self._compute_token_idf(list(self.chunk_tokens_by_id.values()))

    def _lexical_score(
        self,
        query_counter: Counter[str],
        candidate_counter: Counter[str],
        idf_map: dict[str, float],
    ) -> float:
        if not query_counter or not candidate_counter:
            return 0.0
        overlap = 0.0
        candidate_weight = 0.0
        query_weight = 0.0
        for token, count in query_counter.items():
            weight = float(idf_map.get(token, 1.0))
            query_weight += weight * count
            if token in candidate_counter:
                overlap += weight * min(count, candidate_counter[token])
        for token, count in candidate_counter.items():
            candidate_weight += float(idf_map.get(token, 1.0)) * count
        normalizer = max(np.sqrt(query_weight * candidate_weight), 1e-6)
        return float(overlap / normalizer)

    def _build_page_records(self, doc_id: str, page_payloads: Sequence[Dict[str, Any]], text: str) -> List[Dict[str, Any]]:
        records: List[Dict[str, Any]] = []
        cursor = 0
        normalized_pages = [self._normalize_text(str(item.get("text", ""))) for item in page_payloads]
        usable_pages = [item for item in normalized_pages if item]

        if not usable_pages:
            usable_pages = [text]

        for index, page_text in enumerate(usable_pages):
            char_start = cursor
            char_end = char_start + len(page_text)
            records.append(
                {
                    "doc_id": doc_id,
                    "page_number": index + 1,
                    "char_start": char_start,
                    "char_end": char_end,
                    "preview": self._local_summary(page_text, 180),
                    "text": page_text,
                }
            )
            cursor = char_end + 2
        return records

    def _resolve_chunk_pages(self, char_start: int, char_end: int, page_records: Sequence[Dict[str, Any]]) -> tuple[int | None, int | None]:
        overlaps: List[int] = []
        for page in page_records:
            page_start = int(page.get("char_start", 0))
            page_end = int(page.get("char_end", 0))
            if char_end < page_start or char_start > page_end:
                continue
            overlaps.append(int(page.get("page_number", 0)))
        if not overlaps:
            return (page_records[0]["page_number"], page_records[0]["page_number"]) if page_records else (None, None)
        return overlaps[0], overlaps[-1]

    def _normalize_text(self, text: str) -> str:
        text = text.replace("\r", "\n").replace("\xa0", " ")
        text = re.sub(r"[ \t]+", " ", text)
        text = re.sub(r"\n{3,}", "\n\n", text)
        return text.strip()

    def _is_heading(self, line: str) -> bool:
        candidate = line.strip().strip("#").strip()
        if not candidate or len(candidate) > 80:
            return False
        if len(candidate) < 3:
            return False
        if re.search(r"(=|\+|/|\\|\{|\}|\[|\]|\(|\)|∑|Δ|∂|σ|α|β)", candidate):
            return False
        letter_count = len(re.findall(r"[A-Za-z\u4e00-\u9fff]", candidate))
        if letter_count < 2:
            return False
        patterns = [
            r"^第[一二三四五六七八九十百零0-9]+[章节部篇卷回节]",
            r"^(chapter|section|part|unit)\s+[0-9ivx]+",
            r"^[0-9]+(\.[0-9]+){0,3}\s+\S+",
        ]
        if any(re.match(pattern, candidate, re.IGNORECASE) for pattern in patterns):
            return True
        if re.fullmatch(r"[A-Z][A-Z0-9\s\-:,]{6,}", candidate):
            return True
        if re.search(r"[。！？.!?]", candidate):
            return False
        if re.search(r"[\u4e00-\u9fff]", candidate):
            return len(re.findall(r"[\u4e00-\u9fff]", candidate)) >= 4
        words = candidate.split()
        return 4 <= len(candidate) <= 40 and len(words) <= 6 and candidate[:1].isupper()

    def _looks_like_noise_line(self, line: str) -> bool:
        candidate = line.strip()
        if not candidate:
            return True
        if any(re.match(pattern, candidate.lower()) for pattern in NOISE_PATTERNS):
            return True
        if len(candidate) <= 1:
            return True
        letters = len(re.findall(r"[A-Za-z\u4e00-\u9fff]", candidate))
        digits = len(re.findall(r"\d", candidate))
        symbols = len(re.findall(r"[^A-Za-z0-9\u4e00-\u9fff\s]", candidate))
        if letters <= 1 and digits <= 2 and symbols >= 1:
            return True
        if len(candidate) < 10 and letters < 3 and not self._is_heading(candidate):
            return True
        if symbols > 0 and symbols / max(len(candidate), 1) > 0.3 and letters < 10:
            return True
        if re.search(r"(=|\+|/|\\|\{|\}|\[|\]|\(|\)|∑|Δ|∂|σ|α|β)", candidate) and letters < 10:
            return True
        return False

    def _clean_line(self, line: str) -> str:
        candidate = line.strip()
        candidate = re.sub(r"\s+", " ", candidate)
        return candidate.strip("•·|")

    def _sanitize_block(self, text: str) -> str:
        cleaned = re.sub(r"\s+", " ", text).strip()
        return re.sub(r"\s+([,.;:!?])", r"\1", cleaned)

    def _sanitize_table_block(self, text: str) -> str:
        rows = []
        for raw_line in text.splitlines():
            line = re.sub(r"\s{2,}", " | ", raw_line.strip())
            line = re.sub(r"\t+", " | ", line)
            line = re.sub(r"\s+\|\s+", " | ", line)
            if line:
                rows.append(line)
        return "\n".join(rows).strip()

    def _is_table_like_block(self, text: str) -> bool:
        if "\t" in text or "|" in text:
            return True
        if len(re.findall(r"\d+(?:\.\d+)?%?", text)) >= 3 and "  " in text:
            return True
        if len(re.findall(r"(?:\S+\s{2,}\S+)", text)) >= 2:
            return True
        return False

    def _looks_like_table_line(self, line: str) -> bool:
        if "\t" in line or "|" in line:
            return True
        if len(re.findall(r"\d+(?:\.\d+)?%?", line)) >= 2 and "  " in line:
            return True
        return False

    def _valid_block(self, text: str) -> bool:
        cleaned = self._sanitize_block(text)
        if len(cleaned) < 24:
            return False
        letters = len(re.findall(r"[A-Za-z\u4e00-\u9fff]", cleaned))
        symbols = len(re.findall(r"[^A-Za-z0-9\u4e00-\u9fff\s]", cleaned))
        if letters < 10:
            return False
        if symbols / max(len(cleaned), 1) > 0.25 and letters < 18:
            return False
        return True

    def _paragraphs_from_text(self, text: str) -> List[str]:
        normalized = self._normalize_text(text)
        if not normalized:
            return []

        blocks: List[str] = []
        current: List[str] = []
        lines = [self._clean_line(line) for line in normalized.splitlines()]

        def flush_current() -> None:
            nonlocal current
            if not current:
                return
            separator = "\n" if any(self._looks_like_table_line(item) for item in current) else " "
            block = self._sanitize_block(separator.join(current))
            if self._valid_block(block):
                blocks.append(block)
            current = []

        for raw_line in lines:
            line = raw_line.strip()
            if not line:
                flush_current()
                continue

            if self._looks_like_noise_line(line):
                continue

            if self._is_heading(line):
                flush_current()
                blocks.append(f"## {line}")
                continue

            if self._looks_like_table_line(line):
                current.append(line)
                continue

            current.append(line)
            joined = " ".join(current)
            if len(joined) >= 320 and re.search(r"[。！？.!?;；:]$", line):
                flush_current()

        flush_current()
        return blocks

    def _local_summary(self, text: str, limit: int = 260) -> str:
        cleaned = re.sub(r"\s+", " ", text).strip()
        if len(cleaned) <= limit:
            return cleaned
        return cleaned[:limit].rstrip() + "..."

    def _default_category(self, text: str, filename: str) -> str:
        probe = f"{filename}\n{text[:3000]}".lower()
        keyword_map = {
            "金融": ["金融", "股票", "投资", "基金", "银行", "资产", "market", "finance", "trading"],
            "医学": ["医学", "临床", "疾病", "治疗", "患者", "病理", "medical", "health", "drug"],
            "法律": ["法律", "法规", "合同", "司法", "判决", "诉讼", "law", "legal", "compliance"],
            "科技": ["模型", "算法", "系统", "技术", "软件", "rag", "llm", "ai", "research"],
            "生活": ["生活", "健康", "教育", "家庭", "旅行", "饮食", "daily", "lifestyle"],
        }
        for category, keywords in keyword_map.items():
            if any(keyword in probe for keyword in keywords):
                return category
        return "生活"

    def _default_chunk_title(self, text: str, index: int) -> str:
        cleaned = self._sanitize_block(text)
        if not cleaned:
            return f"分段 {index + 1}"
        sentence = re.split(r"[。！？.!?；;:：]", cleaned)[0].strip()
        sentence = re.sub(r"^(and|or|the|a|an)\s+", "", sentence, flags=re.IGNORECASE)
        if len(sentence) > 28:
            sentence = sentence[:28].rstrip()
        return sentence or f"分段 {index + 1}"

    def _extract_keywords(self, text: str, headings: Sequence[str]) -> List[str]:
        combined = " ".join([*headings, text[:2000]])
        chinese_terms = re.findall(r"[\u4e00-\u9fff]{2,8}", combined)
        english_terms = re.findall(r"\b[a-zA-Z][a-zA-Z\-]{3,20}\b", combined.lower())
        stop_words = {"chapter", "section", "figure", "table", "using", "based", "study", "method", "results"}
        counter = Counter()
        for term in chinese_terms:
            counter[term] += 1
        for term in english_terms:
            if term not in stop_words:
                counter[term] += 1
        return [term for term, _ in counter.most_common(8)]

    def _normalize_heading(self, heading: str) -> str:
        cleaned = self._sanitize_block(heading).strip("# ").strip()
        if not cleaned:
            return ""
        if len(cleaned) < 3 or len(cleaned) > 80:
            return ""
        if self._looks_like_noise_line(cleaned):
            return ""
        if re.search(r"(=|\+|/|\\|\{|\}|\[|\]|∑|Δ|∂|σ|α|β)", cleaned):
            return ""
        return cleaned

    def _build_sections(self, text: str, doc_id: str, filename: str) -> tuple[list[Dict[str, Any]], list[Dict[str, Any]], str, list[str]]:
        paragraphs = self._paragraphs_from_text(text)
        sections: List[Dict[str, Any]] = []
        chunks: List[Dict[str, Any]] = []
        headings: List[str] = []

        base_title = Path(filename).stem
        current_section_title = base_title
        current_paragraphs: List[str] = []
        current_order = 0
        current_position = 0

        def flush_section() -> None:
            nonlocal current_paragraphs, current_order, current_position
            if not current_paragraphs:
                return

            filtered_paragraphs = [self._sanitize_block(item) for item in current_paragraphs if self._valid_block(item)]
            current_paragraphs = []
            if not filtered_paragraphs:
                return

            section_id = str(uuid.uuid4())
            title = current_section_title or (base_title if current_order == 0 else f"内容单元 {current_order + 1}")
            joined_text = "\n\n".join(filtered_paragraphs)
            section_summary = self._local_summary(joined_text)
            section_chunks = self._chunk_section(
                doc_id=doc_id,
                section_id=section_id,
                section_title=title,
                paragraphs=filtered_paragraphs,
                start_position=current_position,
                page_records=self.pages_by_doc_id.get(doc_id, []),
            )
            if not section_chunks:
                return
            current_position = section_chunks[-1]["char_end"] + 1
            sections.append(
                {
                    "section_id": section_id,
                    "doc_id": doc_id,
                    "title": title,
                    "order": current_order,
                    "summary": section_summary,
                    "chunk_count": len(section_chunks),
                    "preview": self._local_summary(joined_text, 160),
                }
            )
            chunks.extend(section_chunks)
            current_order += 1

        for block in paragraphs:
            if block.startswith("## "):
                heading = self._normalize_heading(block[3:].strip())
                flush_section()
                current_section_title = heading or f"内容单元 {current_order + 1}"
                if heading:
                    headings.append(heading)
            else:
                current_paragraphs.append(block)

        flush_section()

        if not sections:
            fallback = self._sanitize_block(self._normalize_text(text))
            current_section_title = base_title
            current_paragraphs = [item for item in paragraphs if not item.startswith("## ") and self._valid_block(item)]
            if not current_paragraphs and self._valid_block(fallback):
                current_paragraphs = [fallback]
            flush_section()

        doc_title = headings[0] if headings else base_title
        return sections, chunks, doc_title, headings

    def _chunk_section(
        self,
        doc_id: str,
        section_id: str,
        section_title: str,
        paragraphs: Sequence[str],
        start_position: int,
        page_records: Sequence[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        chunks: List[Dict[str, Any]] = []
        bucket: List[str] = []
        bucket_start = start_position
        cursor = start_position

        def finalize_chunk(text: str, char_start: int) -> None:
            chunk_kind = "table" if self._is_table_like_block(text) else "text"
            cleaned_text = self._sanitize_table_block(text) if chunk_kind == "table" else self._sanitize_block(text)
            if chunk_kind == "text" and not self._valid_block(cleaned_text):
                return
            if chunk_kind == "table" and len(cleaned_text) < 16:
                return
            chunk_index = len(chunks)
            char_end = char_start + len(cleaned_text)
            page_start, page_end = self._resolve_chunk_pages(char_start, char_end, page_records)
            chunks.append(
                {
                    "chunk_id": str(uuid.uuid4()),
                    "doc_id": doc_id,
                    "section_id": section_id,
                    "section_title": section_title,
                    "chunk_index": chunk_index,
                    "chunk_title": self._default_chunk_title(cleaned_text, chunk_index),
                    "chunk_kind": chunk_kind,
                    "text": cleaned_text,
                    "preview": self._local_summary(cleaned_text, 160),
                    "char_start": char_start,
                    "char_end": char_end,
                    "page_start": page_start,
                    "page_end": page_end,
                    "word_count": len(re.findall(r"[A-Za-z0-9]+|[\u4e00-\u9fff]", cleaned_text)),
                }
            )

        def flush_bucket() -> None:
            nonlocal bucket, bucket_start
            if not bucket:
                return
            text = "\n\n".join(bucket).strip()
            finalize_chunk(text, bucket_start)
            bucket = []

        for paragraph in paragraphs:
            cleaned_paragraph = self._sanitize_block(paragraph)
            if not self._valid_block(cleaned_paragraph):
                continue
            if not bucket:
                bucket_start = cursor
            candidate = "\n\n".join(bucket + [cleaned_paragraph]).strip() if bucket else cleaned_paragraph
            if bucket and len(candidate) > self.settings.chunk_size:
                flush_bucket()
                bucket_start = cursor
            bucket.append(cleaned_paragraph)
            cursor += len(cleaned_paragraph) + 2

        if bucket:
            flush_bucket()

        return chunks

    def _doc_search_text(self, doc: Dict[str, Any]) -> str:
        headings = " ".join(doc.get("headings", []))
        keywords = " ".join(doc.get("keywords", []))
        return (
            f"{doc.get('category', '')}\n{doc.get('filename', '')}\n{doc.get('title', '')}\n"
            f"{doc.get('summary', '')}\n{headings}\n{keywords}"
        )

    def _section_search_text(self, section: Dict[str, Any]) -> str:
        doc = self.docs_by_id.get(section["doc_id"], {})
        return (
            f"{doc.get('category', '')}\n{doc.get('title', '')}\n{doc.get('filename', '')}\n"
            f"{section.get('title', '')}\n{section.get('summary', '')}\n{section.get('preview', '')}\n"
            f"{' '.join(doc.get('keywords', []))}"
        )

    def _chunk_search_text(self, chunk: Dict[str, Any]) -> str:
        section = self.sections_by_id.get(chunk["section_id"], {})
        doc = self.docs_by_id.get(chunk["doc_id"], {})
        return (
            f"{doc.get('category', '')}\n{doc.get('title', '')}\n{section.get('title', '')}\n"
            f"{chunk.get('chunk_kind', 'text')}\n{chunk.get('chunk_title', '')}\n{chunk.get('text', '')}"
        )

    def _doc_visible(self, doc: Dict[str, Any], user_id: str | None) -> bool:
        if user_id is None:
            return True
        return doc.get("user_id") == user_id

    def list_files(self, user_id: str | None = None) -> List[Dict[str, Any]]:
        return sorted(
            [doc for doc in self.docs if self._doc_visible(doc, user_id)],
            key=lambda item: item.get("uploaded_at", ""),
            reverse=True,
        )

    def _sorted_sections_for_doc(self, doc_id: str) -> List[Dict[str, Any]]:
        sections = [section for section in self.sections if section["doc_id"] == doc_id]
        return sorted(sections, key=lambda item: item.get("order", 0))

    def _sorted_chunks_for_doc(self, doc_id: str) -> List[Dict[str, Any]]:
        def sort_key(chunk: Dict[str, Any]) -> tuple[int, int]:
            section = self.sections_by_id.get(chunk["section_id"], {})
            return section.get("order", 0), chunk.get("chunk_index", 0)

        chunks = [chunk for chunk in self.chunks if chunk["doc_id"] == doc_id]
        return sorted(chunks, key=sort_key)

    def get_catalog(self, user_id: str | None = None) -> List[Dict[str, Any]]:
        catalog: List[Dict[str, Any]] = []
        for doc in self.list_files(user_id=user_id):
            chunks = [
                {
                    "chunk_id": chunk["chunk_id"],
                    "chunk_title": chunk["chunk_title"],
                    "chunk_kind": chunk.get("chunk_kind", "text"),
                    "section_id": chunk["section_id"],
                    "section_title": chunk["section_title"],
                    "chunk_index": chunk["chunk_index"],
                    "preview": chunk["preview"],
                    "page_start": chunk.get("page_start"),
                    "page_end": chunk.get("page_end"),
                }
                for chunk in self._sorted_chunks_for_doc(doc["doc_id"])
            ]
            catalog.append(
                {
                    "doc_id": doc["doc_id"],
                    "filename": doc["filename"],
                    "category": doc["category"],
                    "title": doc["title"],
                    "summary": doc["summary"],
                    "keywords": doc.get("keywords", []),
                    "chunks": chunks,
                }
            )
        return catalog

    def get_document(self, doc_id: str, user_id: str | None = None) -> Dict[str, Any] | None:
        doc = self.docs_by_id.get(doc_id)
        if not doc or not self._doc_visible(doc, user_id):
            return None

        sections = self._sorted_sections_for_doc(doc_id)
        sorted_chunks = self._sorted_chunks_for_doc(doc_id)
        enriched_sections = []
        flat_chunks = []
        for section in sections:
            section_chunks = [chunk for chunk in sorted_chunks if chunk["section_id"] == section["section_id"]]
            previews = [
                {
                    "chunk_id": chunk["chunk_id"],
                    "chunk_index": chunk["chunk_index"],
                    "chunk_title": chunk["chunk_title"],
                    "chunk_kind": chunk.get("chunk_kind", "text"),
                    "section_id": chunk["section_id"],
                    "section_title": chunk["section_title"],
                    "preview": chunk["preview"],
                    "word_count": chunk["word_count"],
                    "page_start": chunk.get("page_start"),
                    "page_end": chunk.get("page_end"),
                }
                for chunk in section_chunks
            ]
            enriched = dict(section)
            enriched["chunk_count"] = int(section.get("chunk_count") or len(section_chunks))
            enriched["previews"] = previews
            enriched_sections.append(enriched)
            flat_chunks.extend(
                {
                    **preview,
                    "text": chunk["text"],
                    "char_start": chunk["char_start"],
                    "char_end": chunk["char_end"],
                    "page_start": chunk.get("page_start"),
                    "page_end": chunk.get("page_end"),
                    "chunk_kind": chunk.get("chunk_kind", "text"),
                }
                for preview, chunk in zip(previews, section_chunks)
            )

        detail = dict(doc)
        detail["sections"] = enriched_sections
        detail["chunks"] = flat_chunks
        detail["pages"] = self.pages_by_doc_id.get(doc_id, [])
        return detail

    def get_section(self, doc_id: str, section_id: str, user_id: str | None = None) -> Dict[str, Any] | None:
        section = self.sections_by_id.get(section_id)
        doc = self.docs_by_id.get(doc_id)
        if not section or section["doc_id"] != doc_id or not doc or not self._doc_visible(doc, user_id):
            return None

        chunks = [chunk for chunk in self.chunks if chunk["section_id"] == section_id and chunk["doc_id"] == doc_id]
        detail = dict(section)
        detail["chunks"] = sorted(chunks, key=lambda item: item["chunk_index"])
        return detail

    def get_page(self, doc_id: str, page_number: int, user_id: str | None = None) -> Dict[str, Any] | None:
        doc = self.docs_by_id.get(doc_id)
        if not doc or not self._doc_visible(doc, user_id):
            return None
        page = next((item for item in self.pages_by_doc_id.get(doc_id, []) if item["page_number"] == page_number), None)
        if not page:
            return None

        page_chunks = [
            chunk
            for chunk in self._sorted_chunks_for_doc(doc_id)
            if (chunk.get("page_start") or 0) <= page_number <= (chunk.get("page_end") or chunk.get("page_start") or 0)
        ]
        return {
            "doc_id": doc_id,
            "page_number": page["page_number"],
            "char_start": page.get("char_start", 0),
            "char_end": page.get("char_end", 0),
            "preview": page.get("preview", ""),
            "text": page.get("text", ""),
            "chunks": [
                {
                    "chunk_id": chunk["chunk_id"],
                    "chunk_index": chunk["chunk_index"],
                    "chunk_title": chunk["chunk_title"],
                    "chunk_kind": chunk.get("chunk_kind", "text"),
                    "section_id": chunk["section_id"],
                    "section_title": chunk["section_title"],
                    "preview": chunk["preview"],
                    "word_count": chunk["word_count"],
                    "page_start": chunk.get("page_start"),
                    "page_end": chunk.get("page_end"),
                }
                for chunk in page_chunks
            ],
        }

    def prepare_document(self, stored_path: Path, original_name: str, preserve_doc_id: str | None = None) -> Dict[str, Any] | None:
        payload = self._read_document_payload(stored_path)
        text = payload["text"]
        if not text.strip():
            return None

        doc_id = preserve_doc_id or str(uuid.uuid4())
        normalized_pages = [self._normalize_text(str(item.get("text", ""))) for item in payload.get("pages", [])]
        normalized_pages = [item for item in normalized_pages if item]
        page_records = self._build_page_records(doc_id, payload.get("pages", []), self._normalize_text(text))
        cleaned_text = "\n\n".join(normalized_pages)
        cleaned_text = cleaned_text.strip() or self._normalize_text(text)
        if not cleaned_text:
            return None

        self.pages_by_doc_id = getattr(self, "pages_by_doc_id", {})
        self.pages_by_doc_id[doc_id] = page_records
        sections, chunks, title, headings = self._build_sections(cleaned_text, doc_id, original_name)
        if not chunks:
            return None

        uploaded_at = datetime.now(timezone.utc).isoformat()
        combined_preview = "\n".join(section["summary"] for section in sections[:3])
        category = self._default_category(cleaned_text, original_name)
        doc_record = {
            "doc_id": doc_id,
            "filename": original_name,
            "category": category,
            "title": title or Path(original_name).stem,
            "suffix": stored_path.suffix.lower(),
            "uploaded_at": uploaded_at,
            "chunk_count": len(chunks),
            "section_count": len(sections),
            "page_count": len(page_records),
            "summary": self._local_summary(combined_preview or cleaned_text),
            "headings": headings[:20],
            "keywords": self._extract_keywords(cleaned_text, headings),
        }

        return {
            "doc": doc_record,
            "sections": sections,
            "chunks": chunks,
            "pages": page_records,
            "text_excerpt": cleaned_text[: self.settings.metadata_excerpt_chars],
        }

    def apply_metadata(self, prepared: Dict[str, Any], metadata: Dict[str, Any]) -> Dict[str, Any]:
        doc = dict(prepared["doc"])
        sections = [dict(section) for section in prepared["sections"]]
        chunks = [dict(chunk) for chunk in prepared["chunks"]]

        category = metadata.get("category")
        if category in self.settings.allowed_categories:
            doc["category"] = category

        title = str(metadata.get("title", "")).strip()
        if title:
            doc["title"] = title[:120]

        summary = str(metadata.get("summary", "")).strip()
        if summary:
            doc["summary"] = summary[:280]

        keywords = metadata.get("keywords")
        if isinstance(keywords, list):
            normalized_keywords = [str(item).strip()[:24] for item in keywords if str(item).strip()]
            if normalized_keywords:
                doc["keywords"] = normalized_keywords[:8]

        heading_titles = metadata.get("headings")
        if isinstance(heading_titles, list):
            doc["headings"] = [str(item).strip() for item in heading_titles if str(item).strip()][:20]

        chunk_titles = metadata.get("chunk_titles", {})
        if isinstance(chunk_titles, dict):
            for chunk in chunks:
                chunk_title = str(chunk_titles.get(chunk["chunk_id"], "")).strip()
                if chunk_title:
                    chunk["chunk_title"] = chunk_title[:80]

        return {
            "doc": doc,
            "sections": sections,
            "chunks": chunks,
            "text_excerpt": prepared.get("text_excerpt", ""),
        }

    def categories_summary(self, user_id: str | None = None) -> List[Dict[str, Any]]:
        summary: Dict[str, int] = {}
        for doc in self.docs:
            if not self._doc_visible(doc, user_id):
                continue
            summary[doc["category"]] = summary.get(doc["category"], 0) + 1
        return [{"category": key, "count": summary[key]} for key in self.settings.allowed_categories if key in summary]

    def hierarchical_search(
        self,
        queries: Sequence[str],
        *,
        categories: Sequence[str] | None = None,
        doc_ids: Sequence[str] | None = None,
        chunk_ids: Sequence[str] | None = None,
        user_id: str | None = None,
    ) -> tuple[List[SearchResult], List[Dict[str, Any]]]:
        traces: List[Dict[str, Any]] = []

        selected_categories = list(categories) if categories else [item["category"] for item in self.categories_summary(user_id=user_id)]
        for category in selected_categories:
            if category:
                traces.append(
                    {
                        "id": category,
                        "label": category,
                        "score": 1.0,
                        "level": "category",
                        "parent_id": None,
                    }
                )

        docs = self.rank_documents(queries, categories=categories, user_id=user_id, limit=max(self.settings.max_doc_candidates, 8))
        if doc_ids:
            doc_set = set(doc_ids)
            docs = [doc for doc in docs if doc["doc_id"] in doc_set]
            if not docs:
                docs = [doc for doc in self.list_files(user_id=user_id) if doc["doc_id"] in doc_set]

        docs = docs[: self.settings.max_doc_candidates]
        for doc in docs:
            traces.append(
                {
                    "id": doc["doc_id"],
                    "label": doc["title"],
                    "score": float(doc.get("score", 0.0)),
                    "level": "document",
                    "parent_id": doc["category"],
                }
            )

        ranked_chunks = self.rank_chunks(
            queries,
            categories=categories,
            doc_ids=[doc["doc_id"] for doc in docs] if docs else doc_ids,
            chunk_ids=chunk_ids,
            user_id=user_id,
            limit=self.settings.max_chunk_candidates,
        )

        for chunk in ranked_chunks:
            traces.append(
                {
                    "id": chunk.chunk_id,
                    "label": chunk.chunk_title,
                    "score": chunk.score,
                    "level": "chunk",
                    "parent_id": chunk.doc_id,
                }
            )

        return ranked_chunks[: self.settings.top_k], traces

    def get_chunk_candidates_for_docs(self, doc_ids: Sequence[str], limit_per_doc: int = 8) -> List[Dict[str, Any]]:
        candidates: List[Dict[str, Any]] = []
        doc_set = set(doc_ids)
        for doc_id in doc_ids:
            doc_chunks = [chunk for chunk in self.chunks if chunk["doc_id"] == doc_id]
            for chunk in doc_chunks[:limit_per_doc]:
                doc = self.docs_by_id.get(doc_id, {})
                candidates.append(
                    {
                        "chunk_id": chunk["chunk_id"],
                        "doc_id": doc_id,
                        "doc_title": doc.get("title", ""),
                        "category": doc.get("category", "生活"),
                        "section_id": chunk["section_id"],
                        "chunk_title": chunk["chunk_title"],
                        "chunk_index": chunk["chunk_index"],
                        "preview": chunk["preview"],
                        "section_title": chunk["section_title"],
                    }
                )
        return candidates

    def get_chunks_by_ids(self, chunk_ids: Sequence[str]) -> List[SearchResult]:
        results: List[SearchResult] = []
        for chunk_id in chunk_ids:
            chunk = self.chunks_by_id.get(chunk_id)
            if not chunk:
                continue
            doc = self.docs_by_id.get(chunk["doc_id"], {})
            results.append(
                SearchResult(
                    doc_id=chunk["doc_id"],
                    filename=doc.get("filename", ""),
                    category=doc.get("category", "生活"),
                    title=doc.get("title", ""),
                    section_id=chunk["section_id"],
                    section_title=chunk["section_title"],
                    chunk_id=chunk["chunk_id"],
                    chunk_index=chunk["chunk_index"],
                    chunk_title=chunk.get("chunk_title", f"分段 {chunk['chunk_index'] + 1}"),
                    page_start=chunk.get("page_start"),
                    page_end=chunk.get("page_end"),
                    chunk_kind=chunk.get("chunk_kind", "text"),
                    score=1.0,
                    text=chunk["text"],
                )
            )
        return results

    def search(self, query: str, top_k: int | None = None) -> List[SearchResult]:
        results, _ = self.hierarchical_search([query])
        return results[: top_k or self.settings.top_k]

    def _extract_original_name(self, temp_path: Path) -> str:
        name = temp_path.name
        if "--" not in name:
            return name
        return name.split("--", 1)[1]



    def save_prepared_documents(self, prepared_documents: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
        added = []
        for prepared in prepared_documents:
            doc = dict(prepared["doc"])
            doc_id = str(doc.get("doc_id") or "")
            user_id = str(doc.get("user_id") or "")
            if not doc_id or not user_id:
                raise ValueError("prepared document missing doc_id or user_id")
            self._delete_document_rows(doc_id)
            headings = list(doc.get("headings", []))
            keywords = list(doc.get("keywords", []))
            doc_search_text = (
                f"{doc.get('category', '')}\n{doc.get('filename', '')}\n{doc.get('title', '')}\n"
                f"{doc.get('summary', '')}\n{' '.join(headings)}\n{' '.join(keywords)}"
            )
            self.db.execute(
                """
                INSERT INTO documents (doc_id, user_id, filename, stored_path, category, title, suffix, uploaded_at, chunk_count, section_count, page_count, summary, headings_json, keywords_json, embedding_json)
                VALUES (%(doc_id)s, %(user_id)s, %(filename)s, %(stored_path)s, %(category)s, %(title)s, %(suffix)s, %(uploaded_at)s, %(chunk_count)s, %(section_count)s, %(page_count)s, %(summary)s, %(headings_json)s, %(keywords_json)s, %(embedding_json)s)
                """,
                {
                    "doc_id": doc_id,
                    "user_id": user_id,
                    "filename": doc["filename"],
                    "stored_path": str(doc.get("stored_path") or ""),
                    "category": doc["category"],
                    "title": doc["title"],
                    "suffix": doc["suffix"],
                    "uploaded_at": doc["uploaded_at"],
                    "chunk_count": int(doc["chunk_count"]),
                    "section_count": int(doc["section_count"]),
                    "page_count": int(doc["page_count"]),
                    "summary": doc["summary"],
                    "headings_json": json.dumps(headings, ensure_ascii=False),
                    "keywords_json": json.dumps(keywords, ensure_ascii=False),
                    "embedding_json": self._vector_json(self._encode_text(doc_search_text)),
                },
            )
            section_title_by_id: dict[str, str] = {}
            for section in prepared.get("sections", []):
                section_payload = dict(section)
                section_title_by_id[section_payload["section_id"]] = section_payload["title"]
                section_search_text = (
                    f"{doc.get('category', '')}\n{doc.get('title', '')}\n{doc.get('filename', '')}\n"
                    f"{section_payload.get('title', '')}\n{section_payload.get('summary', '')}\n{section_payload.get('preview', '')}\n"
                    f"{' '.join(keywords)}"
                )
                self.db.execute(
                    """
                    INSERT INTO document_sections (section_id, doc_id, user_id, title, order_index, summary, preview, embedding_json)
                    VALUES (%(section_id)s, %(doc_id)s, %(user_id)s, %(title)s, %(order_index)s, %(summary)s, %(preview)s, %(embedding_json)s)
                    """,
                    {
                        "section_id": section_payload["section_id"],
                        "doc_id": doc_id,
                        "user_id": user_id,
                        "title": section_payload["title"],
                        "order_index": int(section_payload.get("order", 0)),
                        "summary": section_payload["summary"],
                        "preview": section_payload["preview"],
                        "embedding_json": self._vector_json(self._encode_text(section_search_text)),
                    },
                )
            for chunk in prepared.get("chunks", []):
                chunk_payload = dict(chunk)
                section_title = chunk_payload.get("section_title") or section_title_by_id.get(chunk_payload["section_id"], "")
                chunk_search_text = (
                    f"{doc.get('category', '')}\n{doc.get('title', '')}\n{section_title}\n"
                    f"{chunk_payload.get('chunk_kind', 'text')}\n{chunk_payload.get('chunk_title', '')}\n{chunk_payload.get('text', '')}"
                )
                self.db.execute(
                    """
                    INSERT INTO document_chunks (chunk_id, doc_id, user_id, section_id, section_title, chunk_index, chunk_title, chunk_kind, text, preview, char_start, char_end, page_start, page_end, word_count, embedding_json)
                    VALUES (%(chunk_id)s, %(doc_id)s, %(user_id)s, %(section_id)s, %(section_title)s, %(chunk_index)s, %(chunk_title)s, %(chunk_kind)s, %(text)s, %(preview)s, %(char_start)s, %(char_end)s, %(page_start)s, %(page_end)s, %(word_count)s, %(embedding_json)s)
                    """,
                    {
                        "chunk_id": chunk_payload["chunk_id"],
                        "doc_id": doc_id,
                        "user_id": user_id,
                        "section_id": chunk_payload["section_id"],
                        "section_title": chunk_payload["section_title"],
                        "chunk_index": int(chunk_payload["chunk_index"]),
                        "chunk_title": chunk_payload["chunk_title"],
                        "chunk_kind": chunk_payload.get("chunk_kind", "text"),
                        "text": chunk_payload["text"],
                        "preview": chunk_payload["preview"],
                        "char_start": int(chunk_payload["char_start"]),
                        "char_end": int(chunk_payload["char_end"]),
                        "page_start": chunk_payload.get("page_start"),
                        "page_end": chunk_payload.get("page_end"),
                        "word_count": int(chunk_payload["word_count"]),
                        "embedding_json": self._vector_json(self._encode_text(chunk_search_text)),
                    },
                )
            for page in prepared.get("pages", []):
                page_payload = dict(page)
                self.db.execute(
                    """
                    INSERT INTO document_pages (doc_id, page_number, char_start, char_end, preview, text)
                    VALUES (%(doc_id)s, %(page_number)s, %(char_start)s, %(char_end)s, %(preview)s, %(text)s)
                    """,
                    {
                        "doc_id": doc_id,
                        "page_number": int(page_payload["page_number"]),
                        "char_start": int(page_payload.get("char_start", 0)),
                        "char_end": int(page_payload.get("char_end", 0)),
                        "preview": page_payload.get("preview", ""),
                        "text": page_payload.get("text", ""),
                    },
                )
            added.append(doc)
        if prepared_documents:
            self._refresh_from_db()
        return added

    def update_document_title(self, doc_id: str, title: str) -> Dict[str, Any] | None:
        doc = self.docs_by_id.get(doc_id)
        normalized_title = self._sanitize_block(title)
        if not doc or not normalized_title:
            return None
        doc["title"] = normalized_title[:120]
        self.db.execute(
            "UPDATE documents SET title = %(title)s, embedding_json = %(embedding_json)s WHERE doc_id = %(doc_id)s",
            {"doc_id": doc_id, "title": doc["title"], "embedding_json": self._vector_json(self._encode_text(self._doc_search_text(doc)))},
        )
        self._refresh_from_db()
        return self.get_document(doc_id)

    def update_chunk_title(self, doc_id: str, chunk_id: str, chunk_title: str) -> Dict[str, Any] | None:
        chunk = self.chunks_by_id.get(chunk_id)
        normalized_title = self._sanitize_block(chunk_title)
        if not chunk or chunk["doc_id"] != doc_id or not normalized_title:
            return None
        chunk["chunk_title"] = normalized_title[:80]
        self.db.execute(
            "UPDATE document_chunks SET chunk_title = %(chunk_title)s, embedding_json = %(embedding_json)s WHERE chunk_id = %(chunk_id)s",
            {"chunk_id": chunk_id, "chunk_title": chunk["chunk_title"], "embedding_json": self._vector_json(self._encode_text(self._chunk_search_text(chunk)))},
        )
        self._refresh_from_db()
        updated = self.chunks_by_id.get(chunk_id)
        if not updated:
            return None
        return {"chunk_id": updated["chunk_id"], "chunk_index": updated["chunk_index"], "chunk_title": updated["chunk_title"], "section_id": updated["section_id"], "section_title": updated["section_title"], "preview": updated["preview"], "word_count": updated["word_count"]}

    def delete_document(self, doc_id: str, user_id: str | None = None) -> bool:
        doc = self.docs_by_id.get(doc_id)
        if not doc or not self._doc_visible(doc, user_id):
            return False
        Path(str(doc.get("stored_path") or "")).unlink(missing_ok=True)
        self._delete_document_rows(doc_id)
        self._refresh_from_db()
        return True

    def delete_all_documents(self) -> None:
        for doc in self.docs:
            Path(str(doc.get("stored_path") or "")).unlink(missing_ok=True)
        self.db.execute("DELETE FROM document_pages")
        self.db.execute("DELETE FROM document_chunks")
        self.db.execute("DELETE FROM document_sections")
        self.db.execute("DELETE FROM documents")
        self._refresh_from_db()

    def _delete_document_rows(self, doc_id: str) -> None:
        self.db.execute("DELETE FROM document_pages WHERE doc_id = %(doc_id)s", {"doc_id": doc_id})
        self.db.execute("DELETE FROM document_chunks WHERE doc_id = %(doc_id)s", {"doc_id": doc_id})
        self.db.execute("DELETE FROM document_sections WHERE doc_id = %(doc_id)s", {"doc_id": doc_id})
        self.db.execute("DELETE FROM documents WHERE doc_id = %(doc_id)s", {"doc_id": doc_id})

    def add_files(self, file_paths: Sequence[Path]) -> Dict[str, List]:
        prepared_documents: List[Dict[str, Any]] = []
        added: List[Dict[str, Any]] = []
        skipped: List[str] = []
        for temp_path in file_paths:
            suffix = temp_path.suffix.lower()
            original_name = self._extract_original_name(temp_path)
            if suffix not in SUPPORTED_SUFFIXES:
                skipped.append(original_name)
                temp_path.unlink(missing_ok=True)
                continue
            doc_id = str(uuid.uuid4())
            destination = self.settings.uploads_dir / f"{doc_id}{suffix}"
            shutil.move(str(temp_path), destination)
            try:
                prepared = self.prepare_document(destination, original_name, preserve_doc_id=doc_id)
            except Exception:
                destination.unlink(missing_ok=True)
                skipped.append(original_name)
                continue
            if prepared is None:
                destination.unlink(missing_ok=True)
                skipped.append(original_name)
                continue
            prepared_documents.append(prepared)
            added.append(prepared["doc"])
        self.save_prepared_documents(prepared_documents)
        return {"added": added, "skipped": skipped}

    def rank_documents(self, queries: Sequence[str], *, categories: Sequence[str] | None = None, user_id: str | None = None, limit: int = 12) -> List[Dict[str, Any]]:
        valid_queries = [query.strip() for query in queries if query.strip()]
        candidates = [doc for doc in self.docs if self._doc_visible(doc, user_id) and (not categories or doc["category"] in categories)]
        if not valid_queries or not candidates:
            return []
        candidate_ids = [doc["doc_id"] for doc in candidates if doc["doc_id"] in self.doc_index_by_id]
        if not candidate_ids:
            return []
        candidate_indexes = np.asarray([self.doc_index_by_id[doc_id] for doc_id in candidate_ids], dtype=np.int32)
        candidate_matrix = self.doc_matrix[candidate_indexes]
        query_vectors = np.vstack([self._encode_text(query) for query in valid_queries]).astype(np.float32)
        vector_scores = candidate_matrix @ query_vectors.T
        query_counters = [Counter(self._tokenize_search_text(query)) for query in valid_queries]
        scores: dict[str, float] = {}
        for row_index, doc_id in enumerate(candidate_ids):
            vector_score = float(np.max(vector_scores[row_index])) if vector_scores.size else 0.0
            lexical_score = max(
                (self._lexical_score(counter, self.doc_tokens_by_id.get(doc_id, Counter()), self.doc_token_idf) for counter in query_counters),
                default=0.0,
            )
            doc = self.docs_by_id.get(doc_id, {})
            title = str(doc.get("title") or "").lower()
            exact_bonus = 0.0
            for query in valid_queries:
                query_lower = query.lower()
                if query_lower and query_lower in title:
                    exact_bonus = max(exact_bonus, 0.2)
            scores[doc_id] = vector_score * 0.68 + lexical_score * 0.24 + exact_bonus
        ranked = sorted(candidates, key=lambda item: scores.get(item["doc_id"], 0.0), reverse=True)
        return [dict(item, score=round(scores.get(item["doc_id"], 0.0), 4)) for item in ranked[:limit]]

    def rank_chunks(self, queries: Sequence[str], *, categories: Sequence[str] | None = None, doc_ids: Sequence[str] | None = None, chunk_ids: Sequence[str] | None = None, user_id: str | None = None, limit: int = 20) -> List[SearchResult]:
        valid_queries = [query.strip() for query in queries if query.strip()]
        if not valid_queries:
            return []
        allowed_doc_ids = set(doc_ids or [])
        allowed_chunk_ids = set(chunk_ids or [])
        category_set = set(categories or [])
        query_vectors = [self._encode_text(query) for query in valid_queries]
        candidates = []
        for chunk in self.chunks:
            doc = self.docs_by_id.get(chunk["doc_id"], {})
            if not self._doc_visible(doc, user_id):
                continue
            if category_set and doc.get("category") not in category_set:
                continue
            if allowed_doc_ids and chunk["doc_id"] not in allowed_doc_ids:
                continue
            if allowed_chunk_ids and chunk["chunk_id"] not in allowed_chunk_ids:
                continue
            candidates.append(chunk)
        candidate_ids = [chunk["chunk_id"] for chunk in candidates if chunk["chunk_id"] in self.chunk_index_by_id]
        if not candidate_ids:
            return []
        candidate_indexes = np.asarray([self.chunk_index_by_id[chunk_id] for chunk_id in candidate_ids], dtype=np.int32)
        candidate_matrix = self.chunk_matrix[candidate_indexes]
        query_matrix = np.vstack(query_vectors).astype(np.float32)
        vector_scores = candidate_matrix @ query_matrix.T
        query_counters = [Counter(self._tokenize_search_text(query)) for query in valid_queries]
        scored = []
        for row_index, chunk_id in enumerate(candidate_ids):
            chunk = self.chunks_by_id.get(chunk_id)
            if not chunk:
                continue
            vector_score = float(np.max(vector_scores[row_index])) if vector_scores.size else 0.0
            lexical_score = max(
                (self._lexical_score(counter, self.chunk_tokens_by_id.get(chunk_id, Counter()), self.chunk_token_idf) for counter in query_counters),
                default=0.0,
            )
            query_phrase_bonus = 0.0
            chunk_text = f"{chunk.get('chunk_title', '')}\n{chunk.get('preview', '')}\n{chunk.get('text', '')}".lower()
            for query in valid_queries:
                query_lower = query.lower()
                if query_lower and query_lower in chunk_text:
                    query_phrase_bonus = max(query_phrase_bonus, 0.18)
            score = vector_score * 0.62 + lexical_score * 0.3 + query_phrase_bonus
            if chunk.get("chunk_kind") == "table":
                score += 0.04
            if score < self.settings.min_retrieval_score:
                continue
            scored.append((chunk, score))
        scored.sort(key=lambda item: item[1], reverse=True)
        results: List[SearchResult] = []
        for chunk, score in scored[:limit]:
            doc = self.docs_by_id.get(chunk["doc_id"], {})
            results.append(SearchResult(doc_id=chunk["doc_id"], filename=doc.get("filename", ""), category=doc.get("category", "??"), title=doc.get("title", ""), section_id=chunk["section_id"], section_title=chunk["section_title"], chunk_id=chunk["chunk_id"], chunk_index=chunk["chunk_index"], chunk_title=chunk.get("chunk_title", f"?? {chunk['chunk_index'] + 1}"), page_start=chunk.get("page_start"), page_end=chunk.get("page_end"), chunk_kind=chunk.get("chunk_kind", "text"), score=round(score, 4), text=chunk["text"]))
        return results

    def _bootstrap_legacy_json_if_needed(self) -> None:
        if int(self.db.scalar("SELECT COUNT(*) AS count FROM documents", default=0) or 0) > 0 or not self.docs_path.exists():
            return
        docs = self._load_json_list(self.docs_path)
        sections = self._load_json_list(self.sections_path)
        chunks = self._load_json_list(self.chunks_path)
        pages = self._load_json_list(self.pages_path)
        if not docs:
            return
        sections_by_doc: dict[str, list[dict[str, Any]]] = {}
        chunks_by_doc: dict[str, list[dict[str, Any]]] = {}
        pages_by_doc: dict[str, list[dict[str, Any]]] = {}
        for section in sections:
            sections_by_doc.setdefault(str(section.get("doc_id") or ""), []).append(section)
        for chunk in chunks:
            chunks_by_doc.setdefault(str(chunk.get("doc_id") or ""), []).append(chunk)
        for page in pages:
            pages_by_doc.setdefault(str(page.get("doc_id") or ""), []).append(page)
        prepared_documents = []
        for doc in docs:
            doc_id = str(doc.get("doc_id") or "")
            if not doc_id:
                continue
            suffix = str(doc.get("suffix") or "")
            prepared_documents.append({"doc": {**doc, "stored_path": str(self.settings.uploads_dir / f"{doc_id}{suffix}"), "user_id": str(doc.get("user_id") or "legacy")}, "sections": sections_by_doc.get(doc_id, []), "chunks": chunks_by_doc.get(doc_id, []), "pages": pages_by_doc.get(doc_id, [])})
        if prepared_documents:
            self.save_prepared_documents(prepared_documents)

    def _load_json_list(self, path: Path) -> list[dict[str, Any]]:
        if not path.exists():
            return []
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return []
        return payload if isinstance(payload, list) else []
